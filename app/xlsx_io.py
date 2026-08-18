"""Excel template download + bulk-upload parsing for governor schedules.

The template mirrors the spreadsheet the interclubs already pass around
(Patty's format, as filled by Paula for the Western Interclub), and the parser
is header-driven so a governor can upload either our template or their own
sheet in that family.
"""
import io
import re
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = [
    ("FIRST DAY (e.g. 2026-09-12)", 22),
    ("LAST DAY (blank = one-day)", 22),
    ("CLUB NAME", 40),
    ("STATE", 9),
    ("CITY (optional)", 18),
    ("OPEN / AMATEUR / PUPPY", 26),
    ("COCKER STAKES? (YES/-)", 22),
    ("WATER TEST (YES/-)", 20),
]

EXAMPLE = ["2026-09-12", "2026-09-13", "EXAMPLE — Timpanogos Hunting Spaniel Club (delete this row)",
           "UT", "Clarkston", "OPEN / AMATEUR / PUPPY", "-", "-"]


REGION_FILLS = {  # canonical legend colors, for the region banner row
    "East": "2F5FA5", "Mid East": "8F1D22", "Mid West": "1F9D5B",
    "Rocky Mountain": "B84390", "West": "A3921E",
}


def build_template(region: str | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (region or "Field Trial") + " Schedule" if region else "Field Trial Schedule"
    banner = ws.cell(row=1, column=1,
                     value=f"REGION: {region}" if region else "REGION: (fill in — East / Mid East / Mid West / Rocky Mountain / West)")
    banner.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    banner.fill = PatternFill("solid", fgColor=REGION_FILLS.get(region or "", "434549"))
    banner.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.append([h for h, _ in HEADERS])
    for i, (_, width) in enumerate(HEADERS, 1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = width
        c = ws.cell(row=2, column=i)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="681E12")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.append(EXAMPLE)
    for cell in ws[3]:
        cell.font = Font(name="Arial", italic=True, color="888888")
    ws.append([])
    note = ws.cell(row=5, column=1,
                   value="Fill one row per trial. Days of the week are added automatically from the dates — "
                         "no need to type them. Delete the gray example row before uploading. "
                         "Dates: 2026-09-12 style, or use Excel date cells.")
    note.font = Font(name="Arial", italic=True, size=9)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(HEADERS))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_export(events, region: str | None, year: int) -> bytes:
    """The region's current events in the same familiar template layout, plus an
    EVENT ID column so an edited re-upload can update events in place."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{region or 'Field Trial'} {year}"[:31]
    banner = ws.cell(row=1, column=1, value=f"REGION: {region}" if region else "REGION: (none — National / other)")
    banner.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    banner.fill = PatternFill("solid", fgColor=REGION_FILLS.get(region or "", "434549"))
    banner.alignment = Alignment(horizontal="center")
    all_headers = [h for h, _ in HEADERS] + ["STATUS (info only)", "EVENT ID (do not change)"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_headers))
    ws.append(all_headers)
    for i, width in enumerate([w for _, w in HEADERS] + [16, 24], 1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = width
        c = ws.cell(row=2, column=i)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="681E12")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for ev in events:
        stakes = " / ".join(s for s, on in [("OPEN", ev["stakes_open"]), ("AMATEUR", ev["stakes_amateur"]),
                                            ("PUPPY", ev["stakes_puppy"])] if on)
        ws.append([
            ev["start_date"][:10],
            ev["end_date"][:10] if ev["end_date"][:10] != ev["start_date"][:10] else "",
            ev["club"] or ev["title"],
            ev["state"], ev["city"],
            stakes, "YES" if ev["stakes_cocker"] else "-", "YES" if ev["water_test"] else "-",
            ev["status"] + (" +hidden" if ev.get("hidden") else ""),
            ev["id"],
        ])
    note = ws.cell(row=ws.max_row + 2, column=1,
                   value="Edit dates/clubs/cities/stakes and re-upload to UPDATE these events in place "
                         "(you'll see every change before it's saved). Leave EVENT ID untouched. "
                         "Rows you add without an ID become new events. To make NEXT year's schedule "
                         "from this one, change the dates and tick 'treat every row as new' when uploading. "
                         "The STATUS column is informational and ignored on upload.")
    note.font = Font(name="Arial", italic=True, size=9)
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(all_headers))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _norm(v):
    return str(v).strip() if v is not None else ""


def _parse_date(v):
    if isinstance(v, (datetime, date)):
        return (v.date() if isinstance(v, datetime) else v).isoformat()
    s = _norm(v)
    if not s or s == "-":
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d %b %Y", "%b %d %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return "ERR"


REGIONS = ["East", "Mid East", "Mid West", "Rocky Mountain", "West"]


def parse_upload(data: bytes):
    """Returns (rows, declared_region, errors). Rows are event-field dicts (no region set)."""
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    header_row = None
    cols = {}
    declared_region = None
    for row in ws.iter_rows(max_row=5, values_only=True):
        for cell in row:
            m = re.match(r"\s*REGION\s*:?\s*(.+)", _norm(cell), re.I)
            if m:
                text = m.group(1).strip()
                for r in sorted(REGIONS, key=len, reverse=True):
                    if re.search(rf"\b{re.escape(r)}\b", text, re.I):
                        declared_region = r
                        break
    # sheet title can also carry it ("West Schedule", "Western Interclub Fall 2026")
    if not declared_region:
        for r in sorted(REGIONS, key=len, reverse=True):
            if re.search(rf"\b{re.escape(r)}(ern)?\b", ws.title, re.I):
                declared_region = r
                break
    for r_i, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
        texts = [_norm(c).upper() for c in row]
        if any("CLUB" in t for t in texts):
            header_row = r_i
            date_cols = []
            for c_i, t in enumerate(texts):
                if not t:
                    continue
                if "EVENT ID" in t or t == "ID":
                    cols["id"] = c_i
                    continue
                if "STATUS" in t:
                    continue  # informational on export, never read back
                if "WEEK" in t and "DAY" not in t:
                    continue
                if "DATE" in t or "DAY (" in t or t.startswith("FIRST DAY") or t.startswith("LAST DAY"):
                    date_cols.append(c_i)
                elif "DAY" in t:
                    continue  # day-of-week column: derived, ignored
                elif "CLUB" in t:
                    cols["club"] = c_i
                elif "CITY" in t:
                    cols["city"] = c_i
                elif "LOCATION" in t or "STATE" in t:
                    cols["state"] = c_i
                elif "COCKER" in t:
                    cols["cocker"] = c_i
                elif "WATER" in t:
                    cols["water"] = c_i
                elif "OPEN" in t or "STAKE" in t or "AMAT" in t:
                    cols["stakes"] = c_i
            if date_cols:
                cols["start"] = date_cols[0]
                cols["end"] = date_cols[1] if len(date_cols) > 1 else date_cols[0]
            break
    if header_row is None or "club" not in cols or "start" not in cols:
        return [], declared_region, ["Could not find a header row with CLUB and DATE columns — is this the right spreadsheet?"]

    rows, errors = [], []
    for r_i, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        get = lambda key: row[cols[key]] if key in cols and cols[key] < len(row) else None
        club = _norm(get("club"))
        if not club or club == "-" or "EXAMPLE" in club.upper():
            continue
        start = _parse_date(get("start"))
        end = _parse_date(get("end"))
        if start in (None, "ERR"):
            errors.append(f"Row {r_i} ({club}): missing or unreadable first-day date — skipped")
            continue
        if end in (None, "ERR"):
            end = start
        if end < start:
            start, end = end, start
        stakes = _norm(get("stakes")).upper()
        yes = lambda v: _norm(v).upper() in ("YES", "Y", "X", "TRUE") or "YES" in _norm(v).upper()
        row_id = None
        if "id" in cols:
            raw_id = get("id")
            try:
                row_id = int(float(raw_id)) if _norm(raw_id) else None
            except (TypeError, ValueError):
                errors.append(f"Row {r_i} ({club}): EVENT ID '{_norm(raw_id)}' isn't a number — treated as a new event")
        rows.append({
            "id": row_id,
            "title": club,
            "club": club,
            "start_date": start,
            "end_date": end,
            "state": _norm(get("state")).upper()[:2],
            "city": _norm(get("city")),
            "stakes_open": 1 if "OPEN" in stakes else 0,
            "stakes_amateur": 1 if "AMAT" in stakes else 0,
            "stakes_puppy": 1 if "PUP" in stakes else 0,
            "stakes_cocker": 1 if ("COCKER" in stakes or yes(get("cocker"))) else 0,
            "water_test": 1 if yes(get("water")) else 0,
            "event_type": "Field Trial",
            "status": "scheduled",
            "source": "xlsx-import",
        })
    return rows, declared_region, errors
